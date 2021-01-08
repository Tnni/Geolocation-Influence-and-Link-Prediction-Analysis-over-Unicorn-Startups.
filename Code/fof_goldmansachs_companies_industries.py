import openpyxl
import networkx as nx
import matplotlib.pyplot as plt
import pickle
from collections import defaultdict
from datetime import datetime
from collections import OrderedDict
import pandas as pd
import numpy as np

wb = openpyxl.load_workbook('../630Data_CB.xlsx')
ws = wb['Sheet1']
industries = defaultdict(list)
investors = set([])


# === filter the dates
def try_parsing_date(text):
    for fmt in ('%b %Y', '%b %d, %Y', '%Y'):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    raise ValueError('no valid date format found')


for row in range(2, ws.max_row+1):
    top_investors = [x.strip() for x in ws[row][8].value.split(',')]
    for t in top_investors:
        if t == 'Goldman Sachs':
            for ind in [x.strip() for x in ws[row][7].value.split(',')]:
                # print(ws[row][10].value)
                comp_date = (ws[row][2].value,
                             try_parsing_date(ws[row][13].value))
                industries[ind].append(comp_date)

# === Sorted for most invested industries from Goldman Sachs
industries = dict(OrderedDict(sorted(industries.items(),
                                     key=lambda item: len(item[1]), reverse=True)))
# === top 5 most invested industries
top_5_industries = list(industries.keys())[:5]
# === Sort each industry by earliest dates.
# this is how the timeline for each industry is created
for ind in top_5_industries:
    print(ind)
    industries[ind].sort(key=lambda x: x[1])
    print(industries[ind])


# ====== Now lets look at the friends
friends_of_Goldman_Sachs = pd.read_pickle(
    "./friends_of_Goldman_Sachs.pkl")
print(friends_of_Goldman_Sachs)


ind_dict = {}
for ind in top_5_industries:
    date = []
    for x, y in industries[ind]:
        date.append(y)
    mean = (np.array(date, dtype='datetime64[s]')
            .view('i8')
            .mean()
            .astype('datetime64[s]'))
    ind_dict[ind] = [mean]
print(ind_dict)


# ===== Creating the datafrane of names, friends, industry avg date
Mobiles = []
Mobiles.append(mean)
Names = []
Names.append('Goldman Sachs')

Friends = []
fr = []
for i, name in friends_of_Goldman_Sachs.iterrows():
    fr.append(name.Name)
Friends.append(fr)

final_dict = {}
# final_dict['mobile'] = Mobiles
final_dict['name'] = Names
final_dict['friends'] = Friends
df = pd.DataFrame(final_dict, columns=['name', 'friends'])
# adding the industry rows
for ind in ind_dict:
    df[ind] = ind_dict[ind]
print(df)


def name_to_date_friend(frname, fof_flag):
    # == find the [industry : [(company name, inv_date),...],...] by 'frname'
    f_industries = defaultdict(list)
    f_investors = set([])

    for row in range(2, ws.max_row+1):
        top_investors = [x.strip() for x in ws[row][8].value.split(',')]
        for t in top_investors:
            if t == frname:
                for ind in [x.strip() for x in ws[row][7].value.split(',')]:
                    # print(ws[row][10].value)
                    if ind in top_5_industries:
                        comp_date = (ws[row][2].value,
                                     try_parsing_date(ws[row][13].value))
                        f_industries[ind].append(comp_date)

    # === Sort by the dates
    for ind in top_5_industries:
        f_industries[ind].sort(key=lambda x: x[1])

    # === Get the avg date from top_5_industries
    for ind in top_5_industries:
        date = []
        for x, y in f_industries[ind]:
            date.append(y)
        f_mean = (np.array(date, dtype='datetime64[s]')
                  .view('i8')
                  .mean()
                  .astype('datetime64[s]'))
        ind_dict[ind].append(f_mean)

    # # === Find of the friends list
    # fr = []
    # for i, frs in friends_of_Goldman_Sachs[friends_of_Goldman_Sachs.Name == frname].iterrows():
    #     if frs.Friends != frname:
    #         fr.append(frs.Friends)
    # Friends.append(fr)

    # === Find of the friends list VER 2
    if(fof_flag == 0):
        for i, frs in friends_of_Goldman_Sachs[friends_of_Goldman_Sachs.Name == frname].iterrows():
            if frname in frs.Friends:
                frs.Friends.remove(frname)
        Friends.append(frs.Friends)
    elif(fof_flag == 1):
        Friends.append([])
    # === adding stuff to the arrays
    Mobiles.append(f_mean)
    Names.append(frname)


for i, row in friends_of_Goldman_Sachs.iterrows():
    name_to_date_friend(row[2], 0)

# final_dict['mobile'] = Mobiles
final_dict['name'] = Names
final_dict['friends'] = Friends
df = pd.DataFrame(final_dict, columns=['name', 'friends'])
for ind in ind_dict:
    df[ind] = ind_dict[ind]

print(df)
# df.to_excel('test.xlsx')


# == FOF
# frnds = df.friends[0]
for row in df.friends[1:]:
    for fof in row:
        if fof not in Names:
            name_to_date_friend(fof, 1)
# for fr in df.friends[0]
# for row in df.friends:
#     print(row)
    # for fof in row.friends:
    #     if row.name != fof:
    #         name_to_date_friend(fof, 1)

# final_dict['mobile'] = Mobiles
final_dict['name'] = Names
final_dict['friends'] = Friends
df = pd.DataFrame(final_dict, columns=['name', 'friends'])
for ind in ind_dict:
    df[ind] = ind_dict[ind]


print(df)
fof_ind_date = df
fof_ind_date.to_pickle("./fof_ind_date_america.pkl")
df.to_excel('test_america.xlsx')
