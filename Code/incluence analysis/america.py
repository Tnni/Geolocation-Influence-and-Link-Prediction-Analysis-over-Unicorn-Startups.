from math import radians, sin, cos, acos, asin, sqrt
import random
from bs4 import BeautifulSoup
import requests
import openpyxl
import networkx as nx
import matplotlib.pyplot as plt
import pickle
from datetime import datetime


import google as gl

from googlesearch import search

import wikipedia

import pandas as pd
import os


from geopy.geocoders import Nominatim

import pandas as pd
import geopandas
import matplotlib.pyplot as plt
from shapely.geometry import Point
import numpy as np
from sklearn.cluster import KMeans
wb = openpyxl.load_workbook('./long_lats.xlsx')
ws = wb['Sheet1']
# geolocator = Nominatim(user_agent="my_app_2")


longs = []
lats = []
names = []
adds = []
table = {}
for row in range(2, ws.max_row+1):
    # location = geolocator.geocode(ws[row][2].value)

    longs.append(ws[row][3].value)
    lats.append(ws[row][4].value)
    names.append(ws[row][1].value)
    adds.append(ws[row][2].value)

table['Longitude'] = longs
table['Latitude'] = lats
table['Name'] = names
table['Address'] = adds

df = pd.DataFrame(
    table, columns=['Name', 'Address', 'Longitude', 'Latitude'])
# df.to_excel("long_lats.xlsx")


# ======== clustering learning =========
# ========= GEOPANDAS
gdf = geopandas.GeoDataFrame(
    df, geometry=geopandas.points_from_xy(df.Longitude, df.Latitude))
print(gdf)
# == Alternative way to do the GeoDatafram
# geometry = [Point(xy) for xy in zip(df.Longitude, df.Latitude)]
# # crs = {'init': 'epsg:4326'}
# gdf = geopandas.GeoDataFrame(df, geometry=geometry)


# == Plotting for testing
# f, ax = plt.subplots(1, figsize=(20, 15))
# ax.set_title("LA Crime", fontsize=40)
# ax.set_axis_off()
# #LA.plot(ax=ax,  edgecolor='grey')
# gdf.plot(ax=ax, color='black')
# plt.show()


# == Extracting X and Y
a = pd.Series(gdf['geometry'].apply(lambda p: p.x))
b = pd.Series(gdf['geometry'].apply(lambda p: p.y))
X = np.column_stack((a, b))
# print(X)


# == Find K, by Elbow Method
# wcss = []
# for i in range(1, 14):
#     kmeans = KMeans(n_clusters=i, init='k-means++', random_state=42)
#     kmeans.fit(X)
#     wcss.append(kmeans.inertia_)
# plt.plot(range(1, 14), wcss)
# plt.title('The Elbow Method to Find the optimal value of K')
# plt.xlabel('Number of clusters')
# plt.ylabel('WCSS')
# plt.show()
# == from above we conclude that K = 3 is optimal


# == Kmeans
kmeans = KMeans(n_clusters=3, init='k-means++', random_state=5,  max_iter=400)
y_kmeans = kmeans.fit_predict(X)
centerz = kmeans.cluster_centers_
k = pd.DataFrame(y_kmeans, columns=['cluster'])
gdf = gdf.join(k)
df = df.join(k)
print(centerz)
print()


# == plot Kmeans
# f, ax = plt.subplots(1, figsize=(20, 15))
# # ax.set_title("Cluster of Investors", fontsize=40)
# # ax.set_axis_off()
# world = geopandas.read_file(geopandas.datasets.get_path('naturalearth_lowres'))
# ax = world.plot(color='lightblue', edgecolor='white', linewidth=0.2)
# ax.set_title("Cluster of Investors", fontsize=40)
# gdf.plot(column='cluster', cmap='Set2', ax=ax, markersize=20)
# # == putting the centerz
# plt.scatter(centerz[:, 0], centerz[:, 1], marker="x",
#             s=150, linewidths=5)
# plt.show()


# == get each cluster data
print("gdf: ", gdf)
# == export gdf with cluster numbers kmeans_cluster.xlsx
# gdf.to_excel("kmeans_cluster.xlsx")

# == seperate geopandas by cluster numbers.
# america = gdf[gdf.cluster == 1]
# america_center = centerz[1]
america = gdf[gdf.cluster == 0]
america_center = centerz[0]
# america = gdf[gdf.cluster == 2]
# america_center = centerz[2]

print("america: ", america)
america = america.reset_index()
x = america['geometry'].map(lambda p: p.x)
y = america['geometry'].map(lambda p: p.y)
pointxy = np.column_stack((x, y))
# # print(radians(114.0543297))
# # for (i, j) in np.column_stack((x, y)):
# #     print(i)
Goldman_Sachs = america[america.Name == "Goldman Sachs"]
x_Goldman_Sachs = Goldman_Sachs['geometry'].map(
    lambda p: p.x).astype(float)
y_Goldman_Sachs = Goldman_Sachs['geometry'].map(
    lambda p: p.y).astype(float)

# print(X)
dist_Goldman_Sachs = []
for i, j in pointxy:
    i = i.astype(float)
    j = j.astype(float)
    # if (i == x_Goldman_Sachs and j == y_Goldman_Sachs):
    #     dist_Goldman_Sachs.append(0)
    # else:
    # formula for the distance
    dist = 6371.01 * acos(sin(y_Goldman_Sachs)*sin(j) +
                          cos(y_Goldman_Sachs)*cos(j)*cos(x_Goldman_Sachs - i))
    dist_Goldman_Sachs.append(dist)

print("DONE")
distdist = np.array(dist_Goldman_Sachs)
kk = pd.DataFrame(distdist, columns=['distance'])
america = america.join(kk)
Goldman_Sachs_row = america.drop(
    america[america.Name != 'Goldman Sachs'].index)
america = america.drop(america[america.Name == 'Goldman Sachs'].index)
print(america)
# america.to_excel("Goldman_Sachs_distance.xlsx")
# ==== finding the average of the
avg_Goldman_Sachs = np.mean(distdist)
print(np.mean(distdist))

# # ==== finding Goldman_Sachs Friends
# frnd = []
# for index, name, address, lon, lat, geo, cluster, dist in america.iterrows():
#     print(dist)


# ==== america updated without Goldman_Sachs =====
friends_of_Goldman_Sachs = america.drop(
    america[america['distance'] > (float(avg_Goldman_Sachs) / 2)].index)
# deleted the distance. Start fresh for Goldman_Sachs friends
del friends_of_Goldman_Sachs['distance']
friends_of_Goldman_Sachs = friends_of_Goldman_Sachs.reset_index()
# friends_of_Goldman_Sachs.to_excel("Goldman_Sachs_friends_below_half_average.xlsx")
# print(friends_of_Goldman_Sachs)

# x = america['geometry'].map(lambda p: p.x).astype(float)
# y = america['geometry'].map(lambda p: p.y).astype(float)
# pointxy = np.column_stack((x, y))

# ======================= DEF ver 1.0 ========================
# def get_distance(framex, originalx):
#     dist_Goldman_Sachs = []
#     for cnt in range(0, len(framex)):
#         # print(frame.iloc[[cnt]]['geometry'])
#         frame = framex.iloc[[cnt]]
#         original = originalx.iloc[[cnt]]
#         x_f = frame['geometry'].map(lambda p: p.x)
#         y_f = frame['geometry'].map(lambda p: p.y)
#         framexy = np.column_stack((x_f, y_f))

#         x = original['geometry'].map(lambda p: p.x).astype(float)
#         y = original['geometry'].map(lambda p: p.y).astype(float)
#         pointxy = np.column_stack((x, y))
#         # for x_f in x_frame:
#         #     print(x_f)
#         # dist_Goldman_Sachs = []
#         for frm in framexy:
#             for pt in pointxy:

#                 # print(type(x_frame) == type(y_frame))
#                 # print(type(x_frame))
#                 # formula for the distance
#                 dist = 6371.01 * acos(sin(frm[1])*sin(pt[1]) +
#                                       cos(frm[1])*cos(pt[1])*cos(frm[0] - pt[0]))
#                 # print(dist)
#             dist_Goldman_Sachs.append(dist)
#             print(dist_Goldman_Sachs)
#     distdist = np.array(dist_Goldman_Sachs)
#     print("distdist: ", distdist)
#     colstr = "distance from " + frame['Name']
#     kk = pd.DataFrame(distdist, columns=[colstr])
#     originalx = originalx.join(kk)
#     return originalx


# america = get_distance(friends_of_Goldman_Sachs, america)
# america.to_excel("Goldman_Sachs_friends_test.xlsx")
# print(america)
# ======================= DEF ver 1.0 ========================
# ======================= DEF ver 2.0 ========================
def get_distance(frame, original):
    dist_Goldman_Sachs = []
    for i, frm in frame.iterrows():
        dist_Goldman_Sachs = []
        for j, og in original.iterrows():
            if (frm[5] == og[4] and frm[4] == og[3]):
                dist_Goldman_Sachs.append(0)
            else:
                # formula for the distance
                # print("frm[5] ", frm[5])
                # print("frm[4] ", frm[4])
                # print("og[4] ", og[4])
                # print("og[3] ", og[3])
                dist = 6371.01 * acos(sin(frm[5])*sin(og[4]) +
                                      cos(frm[5])*cos(og[4])*cos(frm[4] - og[3]))
                dist_Goldman_Sachs.append(dist)

            print("frm[2]: ", frm[2])
            print("og[1]: ", og[1])
            distdist = np.array(dist_Goldman_Sachs)
            # print("distdist: ", distdist)
        print("distdist: ", distdist)
        colstr = "distance from " + frm[2]
        # original = original.join(kk)
        original[colstr] = distdist
        print(original)
    return original
# ======================= DEF ver 2.0 ========================


# ==== test out some smaller tables
# america = america[:12]
# friends_of_Goldman_Sachs = friends_of_Goldman_Sachs[:5]
# print("before going to the functions")
# print("america: ", america)
# print("friends_of_Goldman_Sachs: ", friends_of_Goldman_Sachs)
friends_of_Goldman_Sachs_dist = get_distance(
    friends_of_Goldman_Sachs, america)
# friends_of_Goldman_Sachs_dist.to_excel("Goldman_Sachs_friends_test.xlsx")
print(friends_of_Goldman_Sachs_dist)
# # for i, row in friends_of_Goldman_Sachs.iterrows():
# #     print(row[5])


# ========== find a list of friend of friends based on (below half of average of each)
print("HEREXX")
fof_array = []
for colname in friends_of_Goldman_Sachs_dist.drop(columns=['index', 'Name', 'Address', 'Longitude', 'Latitude', 'geometry', 'cluster', 'distance']):
    # print(row)
    distdist = np.array(friends_of_Goldman_Sachs_dist[colname])
    # distdist = np.array(dist_Goldman_Sachs)
    avg_Goldman_Sachs = np.mean(distdist)
    fof_list = (friends_of_Goldman_Sachs_dist.drop(friends_of_Goldman_Sachs_dist[friends_of_Goldman_Sachs_dist[colname] > (
        float(avg_Goldman_Sachs)/8)].index)['Name']).tolist()
    fof_array.append(fof_list)

Goldman_Sachs_row = Goldman_Sachs_row.drop(columns=['distance'])
print("Goldman_Sachs_row: ", Goldman_Sachs_row)

friends_of_Goldman_Sachs['Friends'] = fof_array
# friends_of_Goldman_Sachs.to_excel(
#     "Goldman_Sachs_friends_of_friend_below_half_average.xlsx")
print("friends of friend : ", friends_of_Goldman_Sachs)
friends_of_Goldman_Sachs.to_pickle(
    "./friends_of_Goldman_Sachs.pkl")
