import openpyxl
import networkx as nx
import matplotlib.pyplot as plt
import pickle
import pandas as pd

with open('company_investor_undirected.pkl', 'rb') as file:
    G = pickle.load(file)

with open('cluster_coloring.pkl', 'rb') as file:
    cluster_coloring = pickle.load(file)

colors = []
for node in G.nodes:
    colors.append(cluster_coloring[node])
    # if cluster_coloring[node] == 0:
    #     colors.append('red')
    # if cluster_coloring[node] == 1:
    #     colors.append('yellow')
    # if cluster_coloring[node] == 2:
    #     colors.append('green')

# nx.draw(G, with_labels=False, edge_color='black', alpha=0.3, node_size=2, width = 0.2, cmap=plt.cm.Reds, node_color='red', colors=cluster_coloring)

wb = openpyxl.load_workbook('./630Data_CB.xlsx')
ws = wb['Sheet1']


data = {}
LOCATION, NO, TOTAL_FUNDING, VALUATION, INDUSTRIES = 0, 1, 4, 5, 7
valid_cols = [LOCATION, NO, TOTAL_FUNDING, VALUATION, INDUSTRIES]
for row in range(2, ws.max_row+1):
    company = ws[row][2].value
    temp = []
    for col in valid_cols:
        temp.append(ws[row][col].value)
    data[company] = temp

for i in range(10):
    count = 0
    industries = []
    for cc in cluster_coloring:
        if cluster_coloring[cc] == i and G.nodes[cc]['type'] == 'company':
            count += 1
            current_industries = [x.strip() for x in data[cc][4].split(',')]
            industries.extend(current_industries)
    print('number of companies in cluster ' + str(i) + ' = ' + str(count))
    if count == 0:
        continue
    temp = pd.Series(industries).value_counts()
    temp = temp[temp>7]
    temp.plot(kind='bar')
    plt.show()