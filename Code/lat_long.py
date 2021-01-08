import random
from bs4 import BeautifulSoup
import requests
import openpyxl
import networkx as nx
import matplotlib.pyplot as plt
import pickle
from datetime import datetime


import google as gl

from googlesearch import search

import wikipedia

import pandas as pd
import os


from geopy.geocoders import Nominatim

import pandas as pd
import geopandas
import matplotlib.pyplot as plt

wb = openpyxl.load_workbook('./final.xlsx')
ws = wb['Sheet1']
geolocator = Nominatim(user_agent="my_app")

all_locs = {}

for row in range(2, ws.max_row+1):
    all_locs[ws[row][2].value] = []

print(all_locs)
print(len(all_locs))

for loc in all_locs:
    location = geolocator.geocode(loc)
    print("location: ", loc)
    all_locs[loc] = [location.longitude, location.latitude]
print(all_locs)
# location = geolocator.geocode(ws[row][2].value)
# print("for ", ws[row][2].value)
# print("longitude: ", location.longitude)
# print("latitude: ", location.latitude)


# geolocator = Nominatim(user_agent="my_app")
longs = []
lats = []
names = []
adds = []
table = {}
for row in range(2, ws.max_row+1):
    # location = geolocator.geocode(ws[row][2].value)

    longs.append(all_locs[ws[row][2].value][0])
    lats.append(all_locs[ws[row][2].value][1])
    names.append(ws[row][1].value)
    adds.append(ws[row][2].value)

table['Longitude'] = longs
table['Latitude'] = lats
table['Name'] = names
table['Address'] = adds

df = pd.DataFrame(table, columns=['Name', 'Address', 'Longitude', 'Latitude'])
# df.to_excel("long_lats.xlsx")

# ========= GEOPANDAS ==========
gdf = geopandas.GeoDataFrame(
    df, geometry=geopandas.points_from_xy(df.Longitude, df.Latitude))

world = geopandas.read_file(geopandas.datasets.get_path('naturalearth_lowres'))

# We restrict to South America.
ax = world[world.continent == 'South America'].plot(
    color='white', edgecolor='black')

# We can now plot our ``GeoDataFrame``.
gdf.plot(ax=ax, color='red')

plt.show()
