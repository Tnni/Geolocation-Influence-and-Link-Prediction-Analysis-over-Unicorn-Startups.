import openpyxl
import networkx as nx
import matplotlib.pyplot as plt
import pickle

wb = openpyxl.load_workbook('./630Data_CB.xlsx')
ws = wb['Sheet1']
company_investor = {}

investors = set([])

for row in range(2, ws.max_row+1):
    top_investors = [x.strip() for x in ws[row][8].value.split(',')]
    for t in top_investors:
        investors.add(t)
    company_investor[ws[row][2].value] = top_investors

G = nx.DiGraph()
for key in company_investor:
    if G.has_node(key):
        G.nodes[key]['type'] = 'company_investor'
    else:
        G.add_node(key, type='company')
    for investor in company_investor[key]:
        if not G.has_node(investor):
            G.add_node(investor, type='investor')
        elif G.nodes[investor]['type'] == 'company':
            G.nodes[investor]['type'] = 'company_investor'
        G.add_edge(investor, key)
c, i = 0, 0
graph_inves = set([])
for n in G.nodes:
    if G.nodes[n]['type'] == 'company' or G.nodes[n]['type'] == 'company_investor':
        c += 1
    if G.nodes[n]['type'] == 'investor' or G.nodes[n]['type'] == 'company_investor':
        i += 1
        graph_inves.add(n)

with open('company_investor_directed.pkl', 'wb') as output:
    pickle.dump(G, output, pickle.HIGHEST_PROTOCOL)
print(c)
print(len(graph_inves))
print(len(company_investor))
print(len(investors))
print(investors - graph_inves)
print(G.number_of_edges())
print(G.number_of_nodes())
# nx.draw(G, with_labels=False, edge_color='black', alpha=0.3, node_size=2, width = 0.2, cmap=plt.cm.Reds, node_color='red')
# plt.show()